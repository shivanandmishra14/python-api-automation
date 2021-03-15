import requests
import json
import openpyxl


# For reading data from excel we need library calle Openpyxl.
# Pip install openpyxl

def test_read_data_from_excel():
    url = "http://thetestingworldapi.com/api/studentsDetails"

    # Open file
    file = open('/Users/b0222643/Documents/PythonAutomationAPI/restApi/Data_Driven_TC/TestDataFIle.json', 'r')

    # Read file and typecast in json format
    json_request = json.loads(file.read())

    # Path of the workbook sheet with openpyxl
    wk = openpyxl.load_workbook('/Users/b0222643/Documents/PythonAutomationAPI/restApi/Data_Driven_TC/Data.xlsx')

    # Now move to sheet
    sh = wk['Sheet1']

    # Reach to rows
    rows = sh.max_rows

    # Start from 2nd row because first is header
    for i in range(2, rows + 1):
        # Read Column in every row
        cell_first_name = sh.cell(row=i, column=1)
        cell_middle_name = sh.cell(row=i, column=2)
        cell_last_name = sh.cell(row=i, column=3)
        cell_date_of_birth = sh.cell(row=i, column=4)

        # These are Four keys in our json
        json_request['first_name'] = cell_first_name.value
        json_request['middle_name'] = cell_middle_name.value
        json_request['last_name'] = cell_last_name.value
        json_request['date_of_birth'] = cell_date_of_birth.value

        response = requests.post(url, json_request)
        print(response.text)
        print(response.status_code)
        assert response.status_code == 201
