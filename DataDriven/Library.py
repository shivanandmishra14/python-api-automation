import json
import jsonpath
import requests
import openpyxl


# Will write common methods so we can reuse it.
# Create a method for row count

class RowCount:

    # Few things are common on both method so create a constructor
    def __init__(self, FileNamePath, SheetName):
        # Make variable global to get used later
        global wk
        global sh

        # Also these two lines are common
        # Path of the workbook sheet with openpyxl and will use the path name as whatever argument we have given
        wk = openpyxl.load_workbook(FileNamePath)

        # Now move to sheet and sheet name will give whatever we have passed on mwthod as an argument
        sh = wk[SheetName]

    # Number of rows
    def fetch_row_count(self):
        # Reach to rows and find number of rows
        rows = sh.max_rows

        # Return rows
        return rows

    # Number of columns
    def fetch_column_count(self):
        col = sh.max_column
        return col

    # Make keys common
    def fetch_key_names(self):
        # Just goto first row  of sheet and find the keys. Fetch how many coloumns we are having in sheet
        c = sh.max_column
        #  Create an empty list
        li = []
        # Run a loop from 1 to 5 , last will get excluded so c+1
        for i in range(1, c + 1):
            cell = sh.cell(row=1, column=i)
            # Whenever we are using insert we need to use two value index and value and index start from zero so i-1
            li.insert(i-1, cell.value)

        return li

    # Update data on JSON
    def update_req_json_data(self, rowNum, jsonRequest, KeyList):
        # FInd the max columns
        c = sh.max_column

        # We are running loop for column because each row we have multiple columns.
        # Create Object for each cell
        for i in range(1, c + 1):
            cell = sh.cell(row=rowNum, column=i)

            # Fetch value of the cell and replace into the json request.
            # So in key list will exclude last because it will start index from 0 and it will till 4.
            jsonRequest[KeyList[i-1]] = cell.value

        return jsonRequest








